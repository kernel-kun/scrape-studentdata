import asyncio
import json
import logging
import os

import httpx
import openpyxl
import pyminizip
from tqdm.asyncio import tqdm_asyncio


# Replace with your desired output file names
ERROR_LOG_FILE = "error_log.txt"
EXCEL_FILE = "student_data.xlsx"
ZIP_FILE = "student_data.zip"
CONCURRENT_REQUESTS = 50  # Number of concurrent requests to be sent to the API

def create_encrypted_zip(files, password):
    """
    Creates an encrypted ZIP file containing the provided files.

    Args:
        files (list): List of file paths to be zipped.
        password (str): Password for encrypting the ZIP file.
    """
    pyminizip.compress_multiple(
        files, [], ZIP_FILE, password, 5
    )  # Adjust compression level as needed

async def get_student_data(roll_no, client, semaphore):
    """
    Sends an async POST request to retrieve student data based on roll number.

    Args:
        roll_no (str): The roll number of the student.
        client (httpx.AsyncClient): The async HTTP client for making requests.
        semaphore (asyncio.Semaphore): Semaphore to limit the number of concurrent requests.

    Returns:
        dict: Dictionary containing student data if successful, otherwise None.
    """
    url = os.environ.get("API_URL")  # Get the API URL from environment variables
    headers = {"Content-Type": "application/x-www-form-urlencoded"}  # Set request headers
    data = {"rollNo": roll_no}  # Payload containing the roll number

    async with semaphore:  # Limit the number of concurrent requests
        try:
            # Send the POST request to the API
            response = await client.post(url, headers=headers, data=data)
            response.raise_for_status()  # Raise an exception for HTTP errors

            if response.content:
                # Decode the response to handle BOM and parse the JSON data
                text = response.text.encode("utf-8").decode("utf-8-sig")
                data = json.loads(text)

                # Check if the rollNo field is empty
                if not data["HTML"].get("rollNo"):
                    logging.warning(f"Empty rollNo for roll number {roll_no}. Skipping...")
                    return None
                else:
                    return data["HTML"]
            else:
                # Handle cases where the response content is empty
                logging.warning(f"Empty response for roll number {roll_no}")
                return None

        except json.JSONDecodeError as json_err:
            # Handle JSON decoding errors
            logging.error(f"JSON decode error for roll number {roll_no}: {json_err}")
            if 'response' in locals():
                logging.error(f"Response content: {response.text}")
            return None

        except httpx.RequestError as e:
            # Handle general request errors, such as connection issues
            logging.error(f"Error for roll number {roll_no}: {e}")
            if 'response' in locals():
                logging.error(f"Response: {response.text}")
            return None

def write_data_to_excel(data, sheet):
    """
    Writes student data to an Excel sheet, dynamically adjusting columns.

    Args:
        data (dict): Dictionary containing student data.
        sheet (openpyxl.worksheet.Worksheet): The Excel sheet to write data to.
    """
    row = sheet.max_row + 1  # Determine the next available row in the sheet

    if row == 2:
        # Write column headers if this is the first row being written
        for col, key in enumerate(data.keys(), start=1):
            sheet.cell(row=1, column=col).value = key

    # Write the student data to the Excel sheet
    for col, (key, value) in enumerate(data.items(), start=1):
        sheet.cell(row=row, column=col).value = value

    logging.info(f"Successfully wrote data for roll number: {data['rollNo']}")

async def main():
    """
    Main function to coordinate fetching data for all roll numbers, writing to Excel,
    and creating an encrypted ZIP file.
    """
    # Configure logging to capture errors and warnings
    logging.basicConfig(
        filename=ERROR_LOG_FILE,
        level=logging.WARN,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )

    # Create a new Excel workbook and set up the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Student Data"

    # Fetch roll numbers from an environment variable and split into a list
    roll_numbers = os.getenv("ROLL_NUMBERS_RANGE").split(" ")

    # Create a semaphore to control the concurrency level
    semaphore = asyncio.Semaphore(CONCURRENT_REQUESTS)
    
    async with httpx.AsyncClient() as client:
        # Create tasks to fetch student data for each roll number
        tasks = [
            get_student_data(str(roll_no), client, semaphore)
            for roll_no in range(int(roll_numbers[0].strip()), int(roll_numbers[1].strip()) + 1)
        ]
        # Execute tasks and display a progress bar using tqdm
        results = await tqdm_asyncio.gather(*tasks, total=len(roll_numbers), desc="Fetching Student Data")

    # Write the fetched data to the Excel file
    for student_data in results:
        if student_data:
            write_data_to_excel(student_data, sheet)

    # Save the Excel workbook
    wb.save(EXCEL_FILE)

    # Create an encrypted ZIP file containing the error log and Excel file
    files_to_zip = [ERROR_LOG_FILE, EXCEL_FILE]
    password = os.environ.get("ENCRYPTION_PASSWORD")
    create_encrypted_zip(files_to_zip, password)

if __name__ == "__main__":
    asyncio.run(main())  # Run the main function asynchronously
