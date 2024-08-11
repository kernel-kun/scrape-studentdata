import json
import logging
import os

import openpyxl
import pyminizip
import requests

# Replace with your desired output file names
ERROR_LOG_FILE = "error_log.txt"
EXCEL_FILE = "student_data.xlsx"
ZIP_FILE = "student_data.zip"


def create_encrypted_zip(files, password):
    pyminizip.compress_multiple(
        files, ZIP_FILE, password, compression=5
    )  # Adjust compression level as needed


def get_student_data(roll_no):
    """
    Sends a POST request to retrieve student data based on roll number.

    Args:
        roll_no (str): The roll number of the student.

    Returns:
        dict: Dictionary containing student data if successful, otherwise None
    """
    # Define API Endpoint, headers and data for the POST request
    url = os.environ.get("API_URL")
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {"rollNo": roll_no}

    try:
        response = requests.post(url, headers=headers, data=data)
        response.raise_for_status()  # Raise an exception for error HTTP status codes
        # Decode the response text using utf-8-sig to handle BOM
        text = response.text.encode("utf-8").decode("utf-8-sig")
        data = json.loads(text)

        # Check if all values are null
        if all(value is None for value in data["HTML"].values()):
            logging.warning(f"All fields are null for roll number {roll_no}")
            return None
        else:
            return data["HTML"]

    except requests.exceptions.RequestException as e:
        logging.error(f"Error for roll number {roll_no}: {e}")
        logging.error(f"Response: {response.text}")
        return None


def write_data_to_excel(data, sheet):
    """
    Writes student data to an Excel sheet, dynamically adjusting columns.

    Args:
        data (dict): Dictionary containing student data.
        sheet (openpyxl.worksheet.Worksheet): The Excel sheet to write data to.
    """
    row = sheet.max_row + 1

    # Get column headers dynamically from the first data row
    if row == 2:
        for col, key in enumerate(data.keys(), start=1):
            sheet.cell(row=1, column=col).value = key

    # Write data to the sheet
    for col, (key, value) in enumerate(data.items(), start=1):
        sheet.cell(row=row, column=col).value = value

    print(f"Successfully wrote data for roll number: {data['rollNo']}")


def main():
    # Configure logging
    logging.basicConfig(
        filename=ERROR_LOG_FILE,
        level=logging.WARN,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )

    # Create a new workbook or open an existing one
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Student Data"

    # Fetch roll numbers from environment variable
    roll_numbers = os.getenv("ROLL_NUMBERS").split("\n")

    for roll_no in roll_numbers:
        student_data = get_student_data(roll_no.strip())
        if student_data:
            write_data_to_excel(student_data, sheet)

    # Save the Excel workbook
    wb.save(EXCEL_FILE)

    # Create an encrypted zip file
    files_to_zip = [ERROR_LOG_FILE, EXCEL_FILE]
    password = os.environ.get("ENCRYPTION_PASSWORD")
    create_encrypted_zip(files_to_zip, password)


if __name__ == "__main__":
    main()
